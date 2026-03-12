"""Experiment package adapter for ProtClaw.

Generates experiment order sheets (Excel) and summary reports (HTML)
for top-ranked protein design candidates.
Input: ranked candidates, project info.
Output: .xlsx order sheet + .html report.
"""

from __future__ import annotations

import json
import os
from datetime import datetime, timezone
from typing import Any

from common.adapter_protocol import BaseTool, ToolResult


class ExperimentPackageAdapter(BaseTool):
    """Adapter for generating experiment order packages."""

    tool_name = "experiment_package"
    tool_version = "1.0.0"

    DEFAULTS: dict[str, Any] = {
        "project_name": "ProtClaw Design",
        "scientist_name": "",
        "top_n": 10,
        "include_report": True,
    }

    def validate_input(self, params: dict[str, Any]) -> dict[str, Any]:
        """Validate experiment package input parameters.

        Required:
            candidates: list[dict] - Ranked candidate data with sequences and scores

        Optional:
            project_name: str - Project name for report header (default: "ProtClaw Design")
            scientist_name: str - Name of the scientist ordering
            top_n: int - Number of top candidates to include (default: 10)
            include_report: bool - Generate HTML report (default: True)
        """
        validated = dict(self.DEFAULTS)
        validated.update(params)

        # Required: candidates
        candidates = validated.get("candidates")
        if not candidates or not isinstance(candidates, list):
            raise ValueError("'candidates' is required and must be a non-empty list")
        for i, c in enumerate(candidates):
            if not isinstance(c, dict):
                raise ValueError(f"candidates[{i}] must be a dict")
            if "sequence" not in c:
                raise ValueError(f"candidates[{i}] must have a 'sequence' field")

        # Validate top_n
        top_n = int(validated["top_n"])
        if top_n < 1:
            raise ValueError(f"top_n must be >= 1, got {top_n}")
        validated["top_n"] = top_n

        validated["project_name"] = str(validated["project_name"])
        validated["scientist_name"] = str(validated["scientist_name"])
        validated["include_report"] = bool(validated["include_report"])

        return validated

    def execute(self, params: dict[str, Any], input_dir: str, output_dir: str) -> ToolResult:
        """Generate experiment order package."""
        candidates = params["candidates"]
        project_name = params["project_name"]
        scientist_name = params["scientist_name"]
        top_n = params["top_n"]
        include_report = params["include_report"]

        # Select top N candidates
        top_candidates = candidates[:top_n]

        output_files: list[str] = []

        # Generate Excel order sheet
        xlsx_path = os.path.join(output_dir, "order_sheet.xlsx")
        _generate_order_sheet(xlsx_path, top_candidates, project_name, scientist_name)
        output_files.append(xlsx_path)

        # Generate HTML report
        if include_report:
            html_path = os.path.join(output_dir, "summary_report.html")
            _generate_html_report(html_path, top_candidates, project_name, scientist_name)
            output_files.append(html_path)

        return self.build_result(
            status="success",
            output_files=output_files,
            metrics={
                "num_candidates_packaged": len(top_candidates),
                "project_name": project_name,
                "files_generated": [os.path.basename(f) for f in output_files],
            },
        )


def _generate_order_sheet(
    output_path: str,
    candidates: list[dict[str, Any]],
    project_name: str,
    scientist_name: str,
) -> None:
    """Generate an Excel order sheet for gene synthesis."""
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Gene Synthesis Order"

        # Header row
        headers = ["Rank", "Name", "Sequence", "Length", "MW (Da)", "pI", "Notes"]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        # Project info
        ws.cell(row=1, column=len(headers) + 2, value="Project:")
        ws.cell(row=1, column=len(headers) + 3, value=project_name)
        ws.cell(row=2, column=len(headers) + 2, value="Scientist:")
        ws.cell(row=2, column=len(headers) + 3, value=scientist_name)
        ws.cell(row=3, column=len(headers) + 2, value="Date:")
        ws.cell(
            row=3,
            column=len(headers) + 3,
            value=datetime.now(tz=timezone.utc).strftime("%Y-%m-%d"),
        )

        # Candidate rows
        for i, candidate in enumerate(candidates, start=1):
            row = i + 1
            ws.cell(row=row, column=1, value=candidate.get("rank", i))
            ws.cell(row=row, column=2, value=candidate.get("name", f"candidate_{i:04d}"))
            ws.cell(row=row, column=3, value=candidate.get("sequence", ""))
            ws.cell(row=row, column=4, value=len(candidate.get("sequence", "")))
            ws.cell(row=row, column=5, value=candidate.get("molecular_weight_da", ""))
            ws.cell(row=row, column=6, value=candidate.get("isoelectric_point", ""))
            ws.cell(row=row, column=7, value=candidate.get("notes", ""))

        wb.save(output_path)

    except ImportError:
        # Fallback: write CSV if openpyxl is not available
        import csv

        csv_path = output_path.replace(".xlsx", ".csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Rank", "Name", "Sequence", "Length", "MW (Da)", "pI", "Notes"])
            for i, candidate in enumerate(candidates, start=1):
                writer.writerow([
                    candidate.get("rank", i),
                    candidate.get("name", f"candidate_{i:04d}"),
                    candidate.get("sequence", ""),
                    len(candidate.get("sequence", "")),
                    candidate.get("molecular_weight_da", ""),
                    candidate.get("isoelectric_point", ""),
                    candidate.get("notes", ""),
                ])

        # Rename to .xlsx path so output_files is consistent
        os.rename(csv_path, output_path)


def _generate_html_report(
    output_path: str,
    candidates: list[dict[str, Any]],
    project_name: str,
    scientist_name: str,
) -> None:
    """Generate an HTML summary report."""
    try:
        import jinja2

        template_str = _HTML_TEMPLATE
        env = jinja2.Environment(autoescape=True)
        template = env.from_string(template_str)
        html = template.render(
            project_name=project_name,
            scientist_name=scientist_name,
            date=datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
            candidates=candidates,
            num_candidates=len(candidates),
        )
        with open(output_path, "w") as f:
            f.write(html)

    except ImportError:
        # Fallback: write a simple HTML without Jinja2
        with open(output_path, "w") as f:
            f.write(f"<html><head><title>{project_name}</title></head><body>\n")
            f.write(f"<h1>{project_name} - Experiment Summary</h1>\n")
            f.write(f"<p>Scientist: {scientist_name}</p>\n")
            f.write(
                f"<p>Date: {datetime.now(tz=timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}</p>\n"
            )
            f.write(f"<p>Candidates: {len(candidates)}</p>\n")
            f.write("<table border='1'><tr><th>Rank</th><th>Name</th>"
                    "<th>Length</th><th>Sequence</th></tr>\n")
            for i, c in enumerate(candidates, start=1):
                seq = c.get("sequence", "")
                display_seq = seq[:50] + "..." if len(seq) > 50 else seq
                f.write(
                    f"<tr><td>{c.get('rank', i)}</td>"
                    f"<td>{c.get('name', f'candidate_{i:04d}')}</td>"
                    f"<td>{len(seq)}</td>"
                    f"<td><code>{display_seq}</code></td></tr>\n"
                )
            f.write("</table></body></html>\n")


_HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
    <title>{{ project_name }} - Experiment Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #2c3e50; }
        .meta { color: #666; margin-bottom: 20px; }
        table { border-collapse: collapse; width: 100%; margin-top: 20px; }
        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
        th { background-color: #3498db; color: white; }
        tr:nth-child(even) { background-color: #f2f2f2; }
        code { font-size: 0.9em; }
    </style>
</head>
<body>
    <h1>{{ project_name }}</h1>
    <div class="meta">
        <p><strong>Scientist:</strong> {{ scientist_name }}</p>
        <p><strong>Date:</strong> {{ date }}</p>
        <p><strong>Candidates:</strong> {{ num_candidates }}</p>
    </div>
    <h2>Top Candidates</h2>
    <table>
        <tr>
            <th>Rank</th>
            <th>Name</th>
            <th>Length</th>
            <th>Sequence (first 50 aa)</th>
        </tr>
        {% for c in candidates %}
        <tr>
            <td>{{ c.get('rank', loop.index) }}</td>
            <td>{{ c.get('name', 'candidate_%04d' % loop.index) }}</td>
            <td>{{ c.get('sequence', '')|length }}</td>
            <td><code>{{ c.get('sequence', '')[:50] }}{% if c.get('sequence', '')|length > 50 %}...{% endif %}</code></td>
        </tr>
        {% endfor %}
    </table>
</body>
</html>"""


def create_adapter() -> ExperimentPackageAdapter:
    """Factory function to create adapter instance."""
    return ExperimentPackageAdapter()
